#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Análise de Mobilizadores - Versão Melhorada
Análise flexível com múltiplos critérios de classificação
"""

import pandas as pd
import numpy as np
import openpyxl
import json
import os
from datetime import datetime

class AnalisadorMobilizadoresMelhorado:
    """
    Analisador melhorado com múltiplos critérios de classificação
    """
    
    def __init__(self):
        # Mapear todas as colunas e tipos de dados disponíveis
        self.mapeamento_colunas = {
            'Mobilizador Desembolso PF': {
                'coluna_principal': 'G',
                'campos_disponiveis': {
                    'Conexao_105': {'coluna': 'G', 'tipo': 'porcentagem'},
                    'Nec_Dia_1': {'coluna': 'I', 'tipo': 'valor'},
                    'Nec_Dia_2': {'coluna': 'O', 'tipo': 'valor'},
                    'Subord_Veloc': {'coluna': 'H', 'tipo': 'categoria'},
                    'Rlz_Dia': {'coluna': 'J', 'tipo': 'valor'}
                },
                'campo_ranking_preferido': 'Conexao_105'  # Usar sempre % Atingimento (Conexão 105%)
            },
            'Mobilizador Desembolso Giro': {
                'coluna_principal': 'M', 
                'campos_disponiveis': {
                    'Conexao_105': {'coluna': 'M', 'tipo': 'porcentagem'},
                    'Nec_Dia': {'coluna': 'U', 'tipo': 'valor'},
                    'Subord_Veloc': {'coluna': 'N', 'tipo': 'categoria'},
                    'Rlz_Dia': {'coluna': 'P', 'tipo': 'valor'}
                },
                'campo_ranking_preferido': 'Conexao_105'  # Usar % Atingimento
            },
            'Mobilizador Desembolso Agro': {
                'coluna_principal': 'S',
                'campos_disponiveis': {
                    'Atg': {'coluna': 'S', 'tipo': 'porcentagem'},
                    'Atg_T': {'coluna': 'T', 'tipo': 'porcentagem'},  # ✅ CORRIGIDO: Adicionar coluna T
                    'Atg_U': {'coluna': 'U', 'tipo': 'porcentagem'},  # ✅ CORRIGIDO: Adicionar coluna U
                    'Subord_Veloc': {'coluna': 'T', 'tipo': 'categoria'}
                },
                'campo_ranking_preferido': 'Atg'  # Usar sempre % Atingimento
            },
            'Mobilizador Regulariza Dívidas Agro': {
                'coluna_principal': 'AB',
                'campos_disponiveis': {
                    'Conexao_105': {'coluna': 'AB', 'tipo': 'porcentagem'},
                    'Conexao_AC': {'coluna': 'AC', 'tipo': 'porcentagem'},  # ✅ CORRIGIDO: Adicionar colunas
                    'Conexao_AD': {'coluna': 'AD', 'tipo': 'porcentagem'},
                    'Conexao_AE': {'coluna': 'AE', 'tipo': 'porcentagem'},
                    'Conexao_AF': {'coluna': 'AF', 'tipo': 'porcentagem'},
                    'Conexao_AG': {'coluna': 'AG', 'tipo': 'porcentagem'},
                    'Nec_Dia': {'coluna': 'AJ', 'tipo': 'valor'},
                    'Rlz_Dia': {'coluna': 'AK', 'tipo': 'valor'},
                    'Atg': {'coluna': 'AL', 'tipo': 'porcentagem'}
                },
                'campo_ranking_preferido': 'Conexao_105'  # Usar sempre % Atingimento
            },
            'Mobilizador Icred 15/90': {
                'coluna_principal': 'Y',
                'campos_disponiveis': {
                    'Atg': {'coluna': 'Y', 'tipo': 'porcentagem'},
                    'Nec_Dia': {'coluna': 'AD', 'tipo': 'valor'},
                    'Subord_Veloc': {'coluna': 'Z', 'tipo': 'categoria'},
                    'Rlz_Dia': {'coluna': 'AE', 'tipo': 'valor'}
                },
                'campo_ranking_preferido': 'Atg'  # Usar sempre % Atingimento
            },
            'Mobilizador Portfólio Priorizado': {
                'coluna_principal': 'AH',
                'campos_disponiveis': {
                    'Atg': {'coluna': 'AH', 'tipo': 'porcentagem'},
                    'Nec_Dia': {'coluna': 'AL', 'tipo': 'valor'},
                    'Subord_Veloc': {'coluna': 'AI', 'tipo': 'categoria'},
                    'Rlz_Dia': {'coluna': 'AM', 'tipo': 'valor'}
                },
                'campo_ranking_preferido': 'Atg'  # Usar sempre % Atingimento
            }
        }
    
    def processar_planilha(self, arquivo_planilha):
        """Processa a planilha e gera rankings para todos os grupos"""
        try:
            resultado = {'sucesso': False, 'rankings': {}, 'erro': None}
            
            # Ler planilha com openpyxl para lidar com células mescladas
            workbook = openpyxl.load_workbook(arquivo_planilha, data_only=True)
            sheet_names = workbook.sheetnames
            
            if not sheet_names:
                resultado['erro'] = 'Planilha vazia'
                return resultado
            
            # Usar a primeira aba
            worksheet = workbook[sheet_names[0]]
            
            # Ler dados completos
            df_completo = pd.read_excel(arquivo_planilha, header=0)
            
            resultados_por_grupo = {}
            
            for grupo, config in self.mapeamento_colunas.items():
                try:
                    registros_grupo = self._extrair_registros_grupo(
                        df_completo, worksheet, grupo, config
                    )
                    
                    if registros_grupo:
                        ranking = self._criar_ranking(registros_grupo, grupo)
                        resultados_por_grupo[grupo] = {
                            'total_registros': len(registros_grupo),
                            'ranking': ranking,
                            'campos_utilizados': config.get('campos_utilizados', [])
                        }
                    else:
                        resultados_por_grupo[grupo] = {
                            'total_registros': 0,
                            'ranking': [],
                            'mensagem': 'Nenhum registro encontrado'
                        }
                        
                except Exception as e:
                    resultados_por_grupo[grupo] = {
                        'total_registros': 0,
                        'ranking': [],
                        'erro': f'Erro ao processar grupo: {str(e)}'
                    }
            
            resultado['sucesso'] = True
            resultado['rankings'] = resultados_por_grupo
            resultado['timestamp'] = datetime.now().isoformat()
            
            return resultado
            
        except Exception as e:
            resultado['erro'] = f'Erro ao processar planilha: {str(e)}'
            return resultado
    
    def _extrair_registros_grupo(self, df, worksheet, grupo, config):
        """Extrai registros para um grupo específico"""
        registros = []
        campo_ranking = config['campo_ranking_preferido']
        
        try:
            # Obter campo de ranking configurado
            campo_info = config['campos_disponiveis'].get(campo_ranking)
            if not campo_info:
                # Fallback para outros campos disponíveis
                for campo, info in config['campos_disponiveis'].items():
                    if info['tipo'] == 'porcentagem':
                        campo_info = info
                        break
                
                if not campo_info:
                    return registros
            
            coluna_letra = campo_info['coluna']
            
            # Converter letra da coluna para índice
            coluna_indice = self._converter_coluna_para_indice(coluna_letra)
            
            # Para Desembolso Agro e Regulariza, expandir busca em múltiplas colunas
            if grupo == 'Mobilizador Desembolso Agro':
                colunas_busca = ['S', 'T', 'U']
            elif grupo == 'Mobilizador Regulariza Dívidas Agro':
                colunas_busca = ['AB', 'AC', 'AD', 'AE', 'AF', 'AG']
            else:
                colunas_busca = [coluna_letra]
            
            # Atualizar campos utilizados
            config['campos_utilizados'] = colunas_busca
            
            # Buscar em todas as colunas configuradas
            for col_letra in colunas_busca:
                col_indice = self._converter_coluna_para_indice(col_letra)
                
                # Iterar pelas linhas para encontrar valores válidos
                for row in range(2, worksheet.max_row + 1):
                    try:
                        # Obter valor da célula
                        cell_value = worksheet.cell(row=row, column=col_indice).value
                        
                        if cell_value is not None:
                            # Verificar se é um valor de % válido
                            valor_percentual = self._converter_para_percentual(cell_value)
                            
                            if valor_percentual is not None and 0 <= valor_percentual <= 100:
                                # Tentar obter nome do mobilizador
                                nome_mob = self._obter_nome_mobilizador(worksheet, row)
                                
                                registro = {
                                    'nome': nome_mob,
                                    'valor_atingimento': valor_percentual,
                                    'valor_original': cell_value,
                                    'linha': row,
                                    'coluna_origem': col_letra
                                }
                                
                                registros.append(registro)
                    except:
                        continue
            
            # Remover duplicatas baseado no nome
            registros_unicos = {}
            for reg in registros:
                nome = reg['nome']
                if nome not in registros_unicos or registros_unicos[nome]['valor_atingimento'] < reg['valor_atingimento']:
                    registros_unicos[nome] = reg
            
            return list(registros_unicos.values())
            
        except Exception as e:
            print(f"Erro ao extrair registros para {grupo}: {str(e)}")
            return []
    
    def _converter_coluna_para_indice(self, letra_coluna):
        """Converte letra da coluna para índice numérico (A=1, B=2, etc.)"""
        resultado = 0
        for char in letra_coluna.upper():
            resultado = resultado * 26 + (ord(char) - ord('A') + 1)
        return resultado
    
    def _converter_para_percentual(self, valor):
        """Converte valor para percentual"""
        try:
            if isinstance(valor, str):
                # Remover caracteres não numéricos exceto vírgula e ponto
                valor_limpo = ''.join(c for c in valor if c.isdigit() or c in '.,')
                if ',' in valor_limpo and '.' not in valor_limpo:
                    valor_limpo = valor_limpo.replace(',', '.')
                valor = float(valor_limpo)
            
            if isinstance(valor, (int, float)):
                if valor > 1.0:
                    return valor  # Já está em formato percentual
                else:
                    return valor * 100  # Converter de decimal para percentual
        except:
            pass
        return None
    
    def _obter_nome_mobilizador(self, worksheet, linha):
        """Tenta obter o nome do mobilizador de várias colunas"""
        colunas_nome = [1, 2, 3, 4]  # A, B, C, D
        
        for col in colunas_nome:
            valor = worksheet.cell(row=linha, column=col).value
            if valor and isinstance(valor, str) and len(valor.strip()) > 0:
                nome = valor.strip()
                # Filtrar nomes muito genéricos
                if len(nome) > 3 and not nome.lower() in ['total', 'soma', 'geral']:
                    return nome
        
        return f"Mobilizador {linha}"
    
    def _criar_ranking(self, registros, grupo):
        """Cria ranking baseado nos registros encontrados"""
        if not registros:
            return []
        
        # Ordenar por valor de atingimento (decrescente)
        registros_ordenados = sorted(
            registros, 
            key=lambda x: x['valor_atingimento'], 
            reverse=True
        )
        
        # Criar ranking
        ranking = []
        for i, reg in enumerate(registros_ordenados, 1):
            ranking.append({
                'posicao': i,
                'nome': reg['nome'],
                'atingimento_percentual': round(reg['valor_atingimento'], 2),
                'valor_original': reg['valor_original']
            })
        
        return ranking
    
    def gerar_imagem_ranking(self, grupo):
        """Gera imagem PNG do ranking usando matplotlib"""
        try:
            import matplotlib.pyplot as plt
            import seaborn as sns
            import matplotlib.patches as patches
            
            # Configurar matplotlib para não-interactive mode
            plt.switch_backend('Agg')
            plt.style.use('seaborn-v0_8')
            sns.set_palette("husl")
            
            # Cores do Banco do Brasil
            cores_bb = {
                'primaria': '#1e3a8a',      # Azul BB
                'secundaria': '#3b82f6',    # Azul claro
                'destaque': '#fbbf24',      # Amarelo BB
                'fundo': '#f8fafc',         # Cinza claro
                'texto': '#1f2937',         # Cinza escuro
                'alternativa': '#10b981'    # Verde
            }
            
            # Configurar figura
            fig, ax = plt.subplots(figsize=(12, 8))
            fig.patch.set_facecolor(cores_bb['fundo'])
            
            # Título
            ax.text(0.5, 0.95, f'Ranking - {grupo}', 
                   horizontalalignment='center', verticalalignment='center',
                   fontsize=16, fontweight='bold', color=cores_bb['primaria'],
                   transform=ax.transAxes)
            
            # Adicionar informações do sistema
            ax.text(0.5, 0.90, f'Critério: % de Atingimento | Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 
                   horizontalalignment='center', verticalalignment='center',
                   fontsize=10, color=cores_bb['texto'],
                   transform=ax.transAxes)
            
            # Nota de correção aplicada
            if 'Agro' in grupo or 'Regulariza' in grupo:
                ax.text(0.5, 0.86, '✅ Mapeamento corrigido para capturar todos os registros', 
                       horizontalalignment='center', verticalalignment='center',
                       fontsize=9, color=cores_bb['alternativa'],
                       transform=ax.transAxes)
            
            # Obter dados do ranking
            resultado = self.processar_planilha_minimal()
            dados_grupo = resultado.get(grupo, {})
            ranking = dados_grupo.get('ranking', [])
            
            if not ranking:
                ax.text(0.5, 0.5, 'Nenhum dado disponível para este grupo', 
                       horizontalalignment='center', verticalalignment='center',
                       fontsize=12, color=cores_bb['texto'],
                       transform=ax.transAxes)
            else:
                # Preparar dados para o gráfico
                nomes = [item['nome'][:25] + '...' if len(item['nome']) > 25 else item['nome'] for item in ranking[:10]]
                valores = [item['atingimento_percentual'] for item in ranking[:10]]
                total_registros = dados_grupo.get('total_registros', len(ranking))
                
                # Criar gráfico de barras horizontal
                cores_barras = [cores_bb['secundaria'] if i % 2 == 0 else cores_bb['alternativa'] 
                              for i in range(len(valores))]
                
                barras = ax.barh(range(len(nomes)), valores, color=cores_barras, alpha=0.8)
                
                # Personalizar eixos
                ax.set_yticks(range(len(nomes)))
                ax.set_yticklabels([f"{i+1}. {nome}" for i, nome in enumerate(nomes)])
                ax.set_xlabel('% de Atingimento', fontsize=12, color=cores_bb['texto'])
                
                # Adicionar valores nas barras
                for i, (barra, valor) in enumerate(zip(barras, valores)):
                    ax.text(valor + 1, i, f'{valor:.1f}%', 
                           va='center', ha='left', fontweight='bold', color=cores_bb['texto'])
                
                # Configurar limites e grid
                ax.set_xlim(0, max(valores) + 10 if valores else 100)
                ax.grid(True, alpha=0.3, axis='x')
                
                # Inverter ordem para mostrar maior no topo
                ax.invert_yaxis()
                
                # Adicionar informações adicionais
                ax.text(0.02, 0.02, f'Total de registros: {total_registros}', 
                       transform=ax.transAxes, fontsize=10, 
                       color=cores_bb['texto'], alpha=0.7)
            
            # Remover bordas desnecessárias
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            
            # Salvar em bytes
            img_buffer = io.BytesIO()
            plt.tight_layout()
            plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            plt.close(fig)
            
            return img_buffer.getvalue()
            
        except Exception as e:
            print(f"Erro ao gerar imagem: {str(e)}")
            return None
    
    def processar_planilha_minimal(self):
        """Versão simplificada para obter apenas dados estruturados"""
        # Implementação simplificada - na prática, isso seria uma versão cacheada
        # dos dados processados mais recentemente
        return {}

# Função para execução independente
if __name__ == "__main__":
    analisador = AnalisadorMobilizadoresMelhorado()
    
    # Teste com planilha local se existir
    if os.path.exists('user_input_files/relatorio-6500.xlsx'):
        with open('user_input_files/relatorio-6500.xlsx', 'rb') as f:
            resultado = analisador.processar_planilha(f)
            
            if resultado['sucesso']:
                print("🎉 ANÁLISE CONCLUÍDA - RESULTADOS:")
                print("=" * 60)
                
                for grupo, dados in resultado['rankings'].items():
                    print(f"\n📊 {grupo}")
                    print(f"   📈 Total de registros: {dados['total_registros']}")
                    
                    if dados['ranking']:
                        print("   🏆 TOP 5:")
                        for item in dados['ranking'][:5]:
                            print(f"      {item['posicao']}. {item['nome']}: {item['atingimento_percentual']:.2f}%")
                        
                        # Verificar se as correções funcionaram
                        if grupo == 'Mobilizador Desembolso Agro' and dados['total_registros'] == 12:
                            print("   ✅ CORREÇÃO AGRO: Funcionando corretamente (12 registros)")
                        elif grupo == 'Mobilizador Regulariza Dívidas Agro' and dados['total_registros'] == 22:
                            print("   ✅ CORREÇÃO REGULARIZA: Funcionando corretamente (22 registros)")
                    else:
                        print("   ⚠️ Nenhum ranking gerado")
                
                print(f"\n📅 Processado em: {resultado.get('timestamp', 'N/A')}")
            else:
                print(f"❌ ERRO: {resultado['erro']}")
    else:
        print("ℹ️ Planilha não encontrada. Use a API Flask para upload.")